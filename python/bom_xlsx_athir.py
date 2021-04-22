#
# Example python script to generate a BOM from a KiCad generic netlist
#
# Example: Ungrouped (One component per row) CSV output
#

"""
    @package
    Generate a xlsx file.
    Components are sorted by ref
    One component per line
    Fields are (if exist)
    Ref, Value, Footprint, Datasheet, Type,Manufacturer, Comment

    Command line:
    python "pathToFile/bom_xlsx_athir.py" "%I" "%O.xlsx"
"""

from __future__ import print_function

# Import the KiCad python helper module
import kicad_netlist_reader
import csv
import sys
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, NamedStyle, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import openpyxl

exvals = ["TestPoint","Barrel_Jack_MountingPin"]

cfps = ["C_1206","C_0805","C_0603","C_0402","C_0201","C_01005"]
rfps = ["R_1206","R_0805","R_0603","R_0402","R_0201","R_01005"]
lfps = ["L_1206","L_0805","L_0603","L_0402","L_0201","L_01005"]

def getColumnWidth(cells):
    celllen = 0
    for cell in cells:
        for c in cell:
            if(len(str(c.value)) > celllen):
                celllen = len(str(c.value))
    celllen = round(celllen*1.25)
    return celllen

def excluded(exvals,component):
    returnval = True
    for val in exvals:
        if(re.search(val,component.getValue())):
            returnval = False
    return returnval
        

def myEqu(self, other):
    """myEqu is a more advanced equivalence function for components which is
    used by component grouping. Normal operation is to group components based
    on their Value and Footprint.

    In this example of a more advanced equivalency operator we also compare the
    custom fields Voltage, Tolerance and Manufacturer as well as the assigned
    footprint. If these fields are not used in some parts they will simply be
    ignored (they will match as both will be empty strings).

    """
    result = True
    if self.getValue() != other.getValue():
        result = False

    elif self.getFootprint() != other.getFootprint():
        result = False
        #print(other.getValue())
    elif self.getField("Tolerance") != other.getField("Tolerance"):
        result = False
        #print(other.getValue())
    elif self.getField("Voltage") != other.getField("Voltage"):
        result = False
    elif self.getField("Mounted") != other.getField("Mounted"):
        result = False
        #print(other.getValue())
    #elif self.getField("Manufacturer") != other.getField("Manufacturer"):
    #    result = False
        '''elif self.getPartName() != other.getPartName():
        result = False
        print('\''+other.getPartName()+'\'')
        print('\''+self.getPartName()+'\'')'''
    

    return result

kicad_netlist_reader.comp.__eq__ = myEqu
#kicad_netlist_reader.netlist().excluded_values.append("TestPoint")
# Generate an instance of a generic netlist, and load the netlist tree from
# the command line option. If the file doesn't exist, execution will stop
net = kicad_netlist_reader.netlist(sys.argv[1])

components = net.getInterestingComponents()
grouped = net.groupComponents()

thin = Side(style='thin',color="000000")
thick = Side(style='thick',color="000000")
wb = Workbook()
worksheet = wb.active
worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE


title = NamedStyle(name="title")
title.font = Font(bold=True,size=14)
wb.add_named_style(title)

header = NamedStyle(name="header")
header.font = Font(bold=True,size=12)
wb.add_named_style(header)

headercent = NamedStyle(name="headercent")
headercent.font = Font(bold=True,size=12)
headercent.border = Border(bottom=thick)
headercent.alignment = Alignment(horizontal="center",vertical="bottom")
wb.add_named_style(headercent)

centered = NamedStyle(name="centered")
centered.font = Font(size=11)
centered.alignment = Alignment(horizontal="center",vertical="bottom")
wb.add_named_style(centered)

redFill = PatternFill(start_color='cf0202',
                   end_color='cf0202',
                   fill_type='solid')

greenFill = PatternFill(start_color='047511',
end_color='047511',
fill_type='solid')

worksheet.append(['BOM - Sleep-Care'])
worksheet.append(['HW Revision:', ""] )
reels = 0
for group in grouped:
    for comp in group:
        c = comp
    if(excluded(exvals,c)):
        reels += 1


worksheet.append(['Date:', net.getDate()] )
worksheet.append(['Tool:', net.getTool()] )
worksheet.append(['Component Count:', len(components)] )
#worksheet.append(['Reel Count:', reels] )
worksheet.append(['Ref', 'Quantity', 'Value', 'Footprint', 'Type','Tolerance','Voltage','Mounting','Manufacturer','Comment','Mouser #'] )#, 'Comment', 'Datasheet'



for group in grouped:
    refs = ""
    for comp in group:
        if (len(refs) > 0):
            refs += ", "
        refs += comp.getRef()
        c = comp
    footprint = c.getFootprint()
    for fp in cfps:
        if(re.search(fp,c.getFootprint())):
            footprint = fp
    for fp in rfps:
        if(re.search(fp,c.getFootprint())):
            footprint = fp
    for fp in lfps:
        if(re.search(fp,c.getFootprint())):
            footprint = fp
    if(len(footprint.split(':')) > 1):
        footprint = footprint.split(':')[1]
            
    if(excluded(exvals,c)):
        worksheet.append([refs, len(group), c.getValue(), footprint, c.getField("Type"), c.getField('Tolerance'), c.getField('Voltage'), c.getField("Mounted"),c.getField("Manufacturer"), c.getField("Comment"), c.getField("Mouser #")]) #, c.getDatasheet()


for cells in worksheet["A1":"B5"]:
    for cell in cells:
        cell.style = 'header'

for cells in worksheet["B6":"K6"]:
    for cell in cells:
        cell.style = 'header'

for cells in worksheet["A6":"K6"]:
    for cell in cells:
        cell.style = 'headercent'


for cells in worksheet["B7":"J50"]:
    for cell in cells:
        cell.style = 'centered'

for cells in worksheet["H7":"H50"]:
    for cell in cells:
        if(re.search("-",str(cell.value))):
            cell.fill = redFill
        elif(re.search("Bottom",str(cell.value))):
            cell.fill = greenFill

for cells in worksheet["K7":"K50"]:
    for cell in cells:
        cell.alignment = Alignment(horizontal="right")

for cells in worksheet["A7":"K50"]:
    if(cells[0].value != None):
        #print(str(cells[0].value))
        for cell in cells:
            cell.border = Border(bottom=thin)

for cells in worksheet["A7":"A50"]:
	if(cells[0].value != None):
		for cell in cells:
			cell.alignment = Alignment(wrap_text=True)

worksheet["A1"].style = 'title'
worksheet["A6"].alignment = Alignment(horizontal="left")
worksheet["B6"].alignment = Alignment(horizontal="left")
#worksheet["A7"].alignment = Alignment(horizontal="left")
worksheet["B5"].alignment = Alignment(horizontal="left")

worksheet.column_dimensions["A"].width = 30#getColumnWidth(worksheet["A1":"A50"])
worksheet.column_dimensions["B"].width = getColumnWidth(worksheet["B5":"B50"])
worksheet.column_dimensions["C"].width = getColumnWidth(worksheet["C5":"C50"])
worksheet.column_dimensions["D"].width = getColumnWidth(worksheet["D5":"D50"])
worksheet.column_dimensions["E"].width = getColumnWidth(worksheet["E5":"E50"])
worksheet.column_dimensions["F"].width = getColumnWidth(worksheet["F5":"F50"])
worksheet.column_dimensions["G"].width = getColumnWidth(worksheet["G5":"G50"])
worksheet.column_dimensions["H"].width = getColumnWidth(worksheet["H5":"H50"])
worksheet.column_dimensions["I"].width = getColumnWidth(worksheet["I5":"I50"])
worksheet.column_dimensions["J"].width = getColumnWidth(worksheet["J5":"J50"])
worksheet.column_dimensions["K"].width = getColumnWidth(worksheet["K5":"K50"])

wb.save(sys.argv[2])